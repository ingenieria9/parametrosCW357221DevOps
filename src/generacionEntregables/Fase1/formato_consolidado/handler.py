import boto3
import json
import re
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook, Workbook
import os

s3 = boto3.client("s3")
TMP_DIR = Path("/tmp")

bucket_name = os.environ["BUCKET_NAME"]

# Carpetas donde buscar
FOLDERS_TO_SEARCH = [
    "files/entregables/Fase1/ACU/PM/",
    "files/entregables/Fase1/ACU/VRP/"
]
OUTPUT_PREFIX = "files/entregables/Fase1/ACU/CIR/"


def lambda_handler(event, context):
    payload = event["payload"]
    cod = payload["COD"]

    # Nombre del archivo consolidado final
    output_filename = f"MPH-EJ-0601-{cod}-F01-ACU-EIN-001.xlsx"
    output_key = f"{OUTPUT_PREFIX}{output_filename}"
    output_path = TMP_DIR / output_filename

    # Crear workbook consolidado
    wb_final = Workbook()
    ws_default = wb_final.active
    ws_default.title = "Resumen"

    # Buscar archivos del COD en los folders definidos
    archivos_cod = []
    for prefix in FOLDERS_TO_SEARCH:
        response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
        if "Contents" not in response:
            continue
        for obj in response["Contents"]:
            key = obj["Key"]
            if key.endswith(".xlsx") and cod in key:
                archivos_cod.append(key)

    if not archivos_cod:
        return {"status": "no_files_found", "cod": cod}

    # Procesar cada archivo encontrado
    for key in archivos_cod:
        print(f"Procesando: {key}")

        # Extraer FID_ELEM del nombre
        match = re.search(
            r"MPH-EJ-0601-[A-Z0-9\-]+-F01-[A-Z]+-EIN-([A-Za-z0-9_]+)\.xlsx",
            key,
        )
        fid_elem = match.group(1) if match else Path(key).stem
        sheet_name = fid_elem[:31]  # Excel máximo 31 caracteres

        # Descargar desde S3 sin guardarlo localmente
        file_stream = BytesIO()
        s3.download_fileobj(bucket_name, key, file_stream)
        file_stream.seek(0)
        wb_src = load_workbook(file_stream)
        ws_src = wb_src.active

        # Crear nueva hoja en consolidado
        ws_dest = wb_final.create_sheet(title=sheet_name)

        # Copiar datos
        for row in ws_src.iter_rows(values_only=True):
            ws_dest.append(row)

    # Eliminar hoja vacía si quedó
    if ws_default.max_row == 1 and ws_default.max_column == 1:
        wb_final.remove(ws_default)

    # Guardar Excel consolidado
    wb_final.save(output_path)

    # Subir a S3
    s3.upload_file(str(output_path), bucket_name, output_key)

    return {
        "status": "ok",
        "cod": cod,
        "total_files": len(archivos_cod),
        "output_file": f"s3://{bucket_name}/{output_key}"
    }
