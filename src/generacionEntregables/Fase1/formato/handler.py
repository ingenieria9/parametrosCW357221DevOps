#Ejemplo de payload 
#{'payload': {'layer_id': 1, 'OBJECTID': '0002', 'geometry': 'null', 'attributes': {'OBJECTID': '0002', 'GlobalID': '3CFDE950-8AE7-440E-B1E7-310C56A35794', 'Identificador': 'PTO_0002', 'Tipo_Punto': 'VRP', 'Creador': 'central_ti_telemetrik', 'Fecha_Creacion': 1758818476306, 'Editor': 'central_ti_telemetrik', 'Fecha_Edicion': 1758829344252, 'Sí': 'Sí', 'Fugas': 'No', 'Signos_de_desgaste': 'null'}, 'point_type': 'VRP'}, 'attachments': ['CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg, CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg, CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg']}

#incoming_payload = event
#incoming_payload = {'payload': {'layer_id': 1, 'OBJECTID': '0002', 'geometry': 'null', 'attributes': {'OBJECTID': '0002', 'GlobalID': '3CFDE950-8AE7-440E-B1E7-310C56A35794', 'Identificador': 'PTO_0002', 'Tipo_Punto': 'VRP', 'Creador': 'central_ti_telemetrik', 'Fecha_Creacion': 1758818476306, 'Editor': 'central_ti_telemetrik', 'Fecha_Edicion': 1758829344252, 'Sí': 'Sí', 'Fugas': 'No', 'Signos_de_desgaste': 'null'}, 'point_type': 'VRP'}, 'attachments': ['CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg, CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg, CW357221-ArcGIS-Data/Puntos/1402_VRP/Fase1/attachment_1402_VRP.jpeg']}

#ejemplo de payload 
'''
{
   "payload":{
      "layer_id":1,
      "OBJECTID":"0002",
      "geometry":"null",
      "attributes":{
         "OBJECTID":"0002",
         "GlobalID":"3CFDE950-8AE7-440E-B1E7-310C56A35794",
         "Creador":"central_ti_telemetrik",
         "Fecha_Creacion":1758818476306,
         "Editor":"central_ti_telemetrik",
         "Fecha_Edicion":1758829344252,
         "id" : "0002",
         "tipo_punto" : "puntos_medicion",
         "signos_desgaste" : "Si",
         "fugas" : "Si",
         "danios" : "No",
         "requiere_instalacion_tapa" : "Si",
         "requiere_limpieza" : "No",
         "razon_limpieza" : "",
         "requiere_clausura" : "",
         "comentario_cond_fisica" : "Tiene una pequeña fuga y desgaste en la tapa",
         "estado_conexion" : "Si",
         "estado_tuberia" : "Si",
         "accesorios_existentes" : "Si",
         "valvula_abre" : "Si",
         "valvula_cierra" : "Si",
         "flujo_agua" : "Si",
         "comentario_conexiones_hid" : "oK",
         "ubicacion_geografica_critica" : "No",
         "posible_expos_fraude" : "No",
         "comentario_vuln" : "Ok",
         "verificacion_4g" : "Si",
         "operador_4g" : "Claro",
         "equipos_usados" : "",
         "conclusiones" : "",
         "recomendaciones" : "Se debe corregir fuga y reemplazar tapa",
         "comentario_general" : "",
         "fecha_modificacion" : "1758818476306",
         "actualizacion_ubicacion" : "No",
         "fecha_creacion" : "1758818476306",
         "latitud" : "37.21",
         "longitud" : "-72.912"
      },
      "point_type":"cajas_medicion"
   },
   "attachments":[
      "files/temp-image-folder/ejemplo1.jpg",
      "files/temp-image-folder/ejemplo2.jpg",
      "files/temp-image-folder/ejemplo3.jpg",
      "files/temp-image-folder/ejemplo4.jpg",
      "files/temp-image-folder/ejemplo5.jpg",
      "files/temp-image-folder/ejemplo6.jpg",
      "files/temp-image-folder/ejemplo7.jpg",
      "files/temp-image-folder/ejemplo8.jpg"
   ]
} 

Variables que vienen de la capa principal y no del payload:
circuito, subcircuito, cuenca, direccion_referencia, vrp
"circuito" : "tmk",
"direccion_referencia" : "Cra 42 #2 cerca al mall",
"vrp" : "vrp-0001",

key s3 de capa principal 
ArcGIS-Data/Puntos/{ID}_{tipo_punto}/Capa_principal/{latest-timestamp}.json
'''

import boto3
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image  # para insertar imágenes
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils import column_index_from_string
import os
from datetime import datetime, timezone, timedelta
import re


s3 = boto3.client("s3")
client_lambda_db = boto3.client("lambda", region_name="us-east-1") 
client_lambda = boto3.client("lambda") 
TMP_DIR = Path("/tmp")

# Parámetros de entrada (variables)

bucket_name = os.environ['BUCKET_NAME']
db_access_arn = os.environ['DB_ACCESS_LAMBDA_ARN']
FORMATO_CONSOLIDADO_LAMBDA_ARN = os.environ['FORMATO_CONSOLIDADO_LAMBDA_ARN']
template_path_s3 = "files/plantillas/Fase1/"
output_path_s3 = "files/entregables/Fase1/"
output_path_s3_for_convert = "files/files-to-convert/Fase1/"


template_name = {"puntos_medicion": "formato-acueducto-pm.xlsx",
                 "vrp": "formato-acueducto-vrp.xlsx", "camara": "formato-alcantarillado.xlsx"}

#MPH-EJ-0601-{CIR_COD}-F01-{ACU/ALC}-EIN-{FID}
COD_name = {"puntos_medicion": "ACU/PM/MPH-EJ-0601-{COD}-F01-ACU-EIN-",
            "vrp": "ACU/VRP/MPH-EJ-0601-{COD}-F01-ACU-EIN-", "camara": "ALC/MPH-EJ-0601-{COD}-F01-ALC-EIN-"}

celdas_imagenes_plantilla = {"puntos_medicion": ["B40", "C40", "D40", "E40","B41", "C41", "D41", "E41", "B42", "C42", "D42", "E42"],
                             "vrp": ["B48", "C48", "D48", "E48","B49", "C49", "D49", "E49", "B50", "C50", "D50", "E50"], "camara": []}


def insert_image(ws, cellNumber, imagen_path):
    img = Image(str(imagen_path))
    cell = ws[cellNumber]

    # Medidas de la celda
    col_width = ws.column_dimensions[cell.column_letter].width   # ancho columna en unidades de Excel
    row_height = ws.row_dimensions[cell.row].height             # alto fila en puntos

    # Conversión aproximada a píxeles
    max_width = col_width * 8
    max_height = row_height * 1.0

    # Escala manteniendo proporciones
    ratio = min(max_width / img.width, max_height / img.height)

    img.width = img.width * ratio
    img.height = img.height * ratio

    ws.add_image(img, cellNumber)

def normalizar_booleans(data_get):
    def convertir_valor(valor):
        if valor is None:  # convertir None a cadena vacía
            return ""
        if isinstance(valor, bool):  # True/False nativos de Python
            return "Si" if valor else "No"
        if isinstance(valor, str):  # "true"/"false" como string
            if valor.lower() == "true":
                return "Si"
            if valor.lower() == "false":
                return "No"
        return valor  # cualquier otro valor queda igual

    # Si es un dict, lo recorremos recursivamente
    if isinstance(data_get, dict):
        return {k: normalizar_booleans(v) for k, v in data_get.items()}
    # Si es una lista, también recursivamente
    if isinstance(data_get, list):
        return [normalizar_booleans(v) for v in data_get]
    
    return convertir_valor(data_get)

def convertir_valores_fecha(data):
    """
    Convierte valores tipo timestamp (en milisegundos) a formato legible
    solo si la clave contiene la palabra 'fecha' (insensible a mayúsculas).
    Interpreta el timestamp como UTC y lo convierte a UTC-5.
    Funciona de forma recursiva para dicts y listas.
    """

    def convertir_fecha(key, valor):
        try:
            if "fecha" in key.lower():
                if isinstance(valor, (int, float)) or (isinstance(valor, str) and valor.isdigit()):
                    timestamp = int(valor) / 1000  # convertir a segundos
                    # Interpretar en UTC y convertir a UTC-5
                    dt_utc = datetime.fromtimestamp(timestamp, tz=timezone.utc)
                    dt_utc_minus_5 = dt_utc.astimezone(timezone(timedelta(hours=-5)))
                    return dt_utc_minus_5.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(f"Error al convertir fecha ({key}): {e}")
        return valor

    if isinstance(data, dict):
        nuevo_dict = {}
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                nuevo_dict[k] = convertir_valores_fecha(v)
            else:
                nuevo_dict[k] = convertir_fecha(k, v)
        return nuevo_dict

    elif isinstance(data, list):
        return [convertir_valores_fecha(v) for v in data]

    else:
        return data

def obtener_consecutivo_s3(bucket, prefix, cod_name):
    """
    Busca el mayor número consecutivo en los nombres de archivos dentro del prefijo dado
    y devuelve el siguiente número disponible (formato 001, 002, etc.).
    """
    max_consec = 0
    pattern = re.escape(cod_name) + r"(\d{3})\.xlsx$"

    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            match = re.search(pattern, key)
            if match:
                num = int(match.group(1))
                if num > max_consec:
                    max_consec = num

    return f"{max_consec + 1:03}"


def obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, CIRCUITO_ACU):
    # Construir el prefijo correcto
    s3_key_capa_principal = (
        f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    )

    # Listar objetos en esa carpeta
    s3_objects = s3.list_objects_v2(Bucket=bucket_name, Prefix=s3_key_capa_principal)

    if "Contents" not in s3_objects:
        print("No hay archivos en la carpeta Capa_principal.")
        return {}

    # Filtrar SOLO archivos que terminen en .json
    json_files = [
        obj for obj in s3_objects["Contents"]
        if obj["Key"].lower().endswith(".json")
    ]

    if not json_files:
        print("No se encontraron archivos .json en Capa_principal.")
        print(f"Se buscó usando Prefix: {s3_key_capa_principal}")
        return {}

    # Tomar el más reciente por fecha
    latest_json = max(json_files, key=lambda x: x["LastModified"])["Key"]

    print(f"Usando archivo principal: {latest_json}")

    # Descargar temporalmente en /tmp
    tmp_path = TMP_DIR / "capa_principal.json"
    s3.download_file(bucket_name, latest_json, str(tmp_path))

    # Leer el contenido con validación
    with open(tmp_path, "r", encoding="utf-8") as f:
        contenido = f.read().strip()
        if not contenido:
            print(" El archivo JSON está vacío.")
            return {}
        try:
            return json.loads(contenido)
        except json.JSONDecodeError as e:
            print(f" Error al parsear JSON {latest_json}: {e}")
            return {}
        
def invoke_lambda_db(payload, FunctionName):
    response = client_lambda_db.invoke(
        FunctionName=FunctionName,
        InvocationType='RequestResponse',
        Payload=json.dumps(payload).encode('utf-8')
    )
    
    # Lee el cuerpo de la respuesta
    result = response["Payload"].read().decode("utf-8")
    
    # Intenta parsear a JSON si es posible
    try:
        return json.loads(result)
    except json.JSONDecodeError:
        return {"raw_response": result}      

def invoke_lambda(payload, FunctionName):
    response = client_lambda.invoke(
        FunctionName=FunctionName,
        InvocationType='Event',  # async
        Payload=json.dumps(payload).encode('utf-8')  #  convierte a JSON y luego a bytes
    )
    return response  

def lambda_handler(event, context):

    payload_data = event["payload"]
    tipo_punto = event["payload"]["TIPO_PUNTO"]
    FID_ELEM = event["payload"]["FID_ELEM"]
    GlobalID = event["payload"]["PARENT_ID"] #id uuid global 
    CIRCUITO_ACU = event["payload"]["CIRCUITO_ACU"].replace(" ", "_")


    #template_path_s3 + devolver de template key el value segun tipo de punto (ej para caja de medicion devuelve formato-acueducto.xlsx)
    template_key = template_path_s3 + template_name.get(tipo_punto)

    # Paths locales en Lambda (/tmp)
    template_path = TMP_DIR / "plantilla.xlsx"
    imagen_keys = event["attachments"]
    # otras imagenes extra de capa principal
    folder_imagen_extra = f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    # Listar objetos PNG en el folder del bucket
    response = s3.list_objects_v2(
        Bucket=bucket_name,
        Prefix=folder_imagen_extra
    )
    # Agregar todos los archivos .png encontrados
    if "Contents" in response:
        for obj in response["Contents"]:
            key = obj["Key"]
            if key.lower().endswith((".png", ".jpeg", ".jpg")):
                imagen_keys.append(key)
    # Eliminar duplicados si existen
    imagen_keys = list(set(imagen_keys))

    imagen_paths = [TMP_DIR / Path(k).name for k in imagen_keys]
    output_path = TMP_DIR / "output.xlsx"

    # Descargar archivos desde S3
    s3.download_file(bucket_name, template_key, str(template_path))
    #s3.download_file(bucket_name, json_key, str(json_path))
    
    #descargar imagenes
    for key, path in zip(imagen_keys, imagen_paths):
        s3.download_file(bucket_name, key, str(path))

    # Cargar plantilla Excel
    wb = load_workbook(template_path)
    ws = wb.active  # hoja específica con wb["NombreHoja"] o activa con wb.active

    # ajustar ancho columnas
    for col in ['B', 'C', 'D', 'E']:
        #print(f"Columna {col}: {ws.column_dimensions[col].width}")
        ws.column_dimensions[col].width = 40
        #print(f"Columna {col}: {ws.column_dimensions[col].width}")
    
    # Leer datos adicionales desde S3
    capa_principal_data = obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, CIRCUITO_ACU)
    #capa_principal_atributos = capa_principal_data["attributes"]
    # Unir ambos diccionarios (payload tiene prioridad si hay claves iguales)
    combined_data = {**capa_principal_data, **payload_data}

    #  Normalizar
    json_data = normalizar_booleans(combined_data)
    json_data = convertir_valores_fecha(json_data)

    # Campos para placeholders
    campos_contexto = list(combined_data.keys())

    # Construir contexto final
    context = {
        f"{{{{{campo}}}}}": json_data.get(campo, "")
        for campo in campos_contexto
    }

    print(context)


    # Reemplazar variables en todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in context.items():
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(value))

    # Insertar imágenes (en celdas específicas)
    #celdas_imagenes = ["B40", "C40", "D40", "E40","B41", "C41", "D41", "E41", "B42", "C42", "D42", "E42"]
    celdas_imagenes = celdas_imagenes_plantilla.get(tipo_punto)

    #Ciclo para insertar las imagenes en las celdas disponibles
    for celda, imagen_path in zip(celdas_imagenes, imagen_paths):
        insert_image(ws, celda, imagen_path)

    # Guardar archivo final
    wb.save(output_path)


    # Subir resultado a S3
    #listar los archivos de s3 en
    # Obtener el consecutivo siguiente en esa carpeta
    #consecutivo = obtener_consecutivo_s3(bucket_name, output_path_s3, COD_name[tipo_punto])

    #obtener de S3 el archivo json que contiene el codigo del circuito para construir el archivo
    # Descargar temporalmente en /tmp
    tmp_path_code = TMP_DIR / "code.json"
    if tipo_punto == "camara":
        code_file = "files/epm_codes/CODE_ALC_CUE.json"
    else:
        code_file = "files/epm_codes/CODE_ACU_CIR.json"
    s3.download_file(bucket_name, code_file, str(tmp_path_code))

    # Leer el contenido con validación
    with open(tmp_path_code, "r", encoding="utf-8") as f:
        contenido = f.read().strip()
        if not contenido:
            print(" El archivo JSON está vacío.")
        try:
            code_json =  json.loads(contenido)
        except json.JSONDecodeError as e:
            print(f" Error al parsear JSON {code_file}: {e}")
    
    code_data = code_json[event["payload"]["CIRCUITO_ACU"]]

    if not code_data:
        code_data = CIRCUITO_ACU

    file_name = COD_name[tipo_punto].format(COD=code_data)

    #Construir el nombre completo del archivo
    output_key = f"{output_path_s3}{file_name}{FID_ELEM}.xlsx"
    convert_output_key = f"{output_path_s3_for_convert}{file_name}{FID_ELEM}.xlsx"

    #Subir a carpeta entregables y a files_to_convert
    s3.upload_file(str(output_path), bucket_name, convert_output_key)
    s3.upload_file(str(output_path), bucket_name, output_key)


    payload_db = {
        "queryStringParameters": {
            "query": f"""SELECT CASE WHEN COUNT(*) = (SELECT COUNT(*)  FROM puntos_capa_principal p2 WHERE p2."CIRCUITO_ACU" = p1."CIRCUITO_ACU"  AND p2."PUNTO_EXISTENTE" = 'Si' AND p2."FASE_INICIAL" = 'fase1'  AND p2."FID_ELEM" IN (SELECT "FID_ELEM" FROM fase_1))THEN 'Finalizado' ELSE 'Incompleto' END AS estado, p1."CIRCUITO_ACU" as "CIRCUITO_ACU" FROM puntos_capa_principal p1 WHERE p1."CIRCUITO_ACU" = ( SELECT "CIRCUITO_ACU" FROM puntos_capa_principal WHERE "GlobalID"  = '{GlobalID}')GROUP BY p1."CIRCUITO_ACU";""",
            "time_column": "fecha_creacion",
            "db_name": "parametros"
        }
    }
    print(payload_db)
    response_db =invoke_lambda_db(payload_db, db_access_arn)
    print(response_db)
    #Parsear el body 
    body = json.loads(response_db["body"])
    # Extraer el valor del campo "estado"
    estado = body[0]["estado"]
    circuito = body[0]["CIRCUITO_ACU"]

    print(estado)
    print(circuito)

    incoming_payload = { "payload": { "COD": file_name } }

    # Si es ultimo punto, invocar a lambda de generación de informe (async)
    if estado == "Finalizado":
        invoke_lambda(incoming_payload, FORMATO_CONSOLIDADO_LAMBDA_ARN)
        print("Invocada lambda formato consolidado")    

    return {
        "status": "ok",
        "output_file": f"s3://{bucket_name}/{output_key}"
    }