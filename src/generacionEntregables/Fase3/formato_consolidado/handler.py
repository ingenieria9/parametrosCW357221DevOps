import boto3
import json
import os
import re
import tempfile
from pathlib import Path
from botocore.exceptions import ClientError
from io import BytesIO
from datetime import datetime, timezone, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

from PIL import Image as PImage, ExifTags


s3 = boto3.client("s3")
TMP_DIR = Path("/tmp")

bucket_name = os.environ["BUCKET_NAME"]

# Rutas conocidas de Lambda A
template_path_s3 = "files/plantillas/Fase3/"
output_path_s3 = "files/entregables/Fase3/"

template_name = {"puntos_medicion": "formato-acueducto-pm.xlsx",
                 "vrp-caudal-PLUM": "formato-acueducto-vrp-caudal-PLUM.xlsx",
                 "vrp-presion_caudal-PLUM": "formato-acueducto-vrp-presion_caudal-PLUM.xlsx",
                 "vrp-presion-Additel": "formato-acueducto-vrp-presion-Additel.xlsx",
                  "vrp-presion-PLUM":  "formato-acueducto-vrp-presion-PLUM.xlsx",
                   "camara": "formato-alcantarillado.xlsx"}

template_general = 'listado-senales'

celdas_imagenes_plantilla = {
    "vrp-presion-Additel": {
        "inst": ["B33", "C33", "D33", "B34", "C34", "D34", "B35", "C35", "D35"],
        "desc": ["B37", "C37", "D37", "B38", "C38", "D38", "B39", "C39", "D39"]
    },
    "vrp-presion-PLUM": {
        "inst": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"],
        "desc": ["B33", "C33", "D33",  "B34", "C34", "D34", "B35", "C35", "D35"]
    },
    "vrp-presion_caudal-PLUM": {
        "inst": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"],
        "desc": ["B33", "C33", "D33"  "B34", "C34", "D34", "B35", "C35", "D35"]
    },
    "vrp-caudal-PLUM": {
        "inst": ["B25", "C25", "D25", "B26", "C26", "D26", "B27", "C27", "D27"],
        "desc": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"]
    },
    "puntos_medicion": {
        "inst": ["B26", "C26", "D26", "B27", "C27", "D27", "B28", "C28", "D28"],
        "desc": ["B30", "C30", "D30", "B31", "C31", "D31", "B32", "C32", "D32"]
    },
    "camara": { #to be defined
        "inst": ["B24", "C24", "D24"],
        "desc": ["B28", "C28", "D28"]
    }
}

datalogger_traduccion = {
"Additel" : "Additel 680A", "PLUM" : "PLUM MacREJ5"
}

variables_a_medir_traduccion = {
    "presion_caudal": "Presión, Caudal","presion": "Presión",
    "caudal": "Caudal","area_velocidad": "Área, velocidad"}

# === Funciones auxiliares ===

def obtener_fecha_exif(image_source):
    try:
        if isinstance(image_source, (str, Path)):
            img = PImage.open(str(image_source))
        else:
            # BytesIO
            image_source.seek(0)
            img = PImage.open(image_source)

        exif_data = img._getexif()
        if not exif_data:
            return None

        exif = {
            ExifTags.TAGS.get(tag, tag): value
            for tag, value in exif_data.items()
        }

        dt = exif.get("DateTimeOriginal") or exif.get("DateTime")
        if not dt:
            return None

        return datetime.strptime(dt, "%Y:%m:%d %H:%M:%S")

    except Exception:
        return None

def clasificar_imagenes_por_fecha(imagen_sources):
    imagenes_con_fecha = []

    for src in imagen_sources:
        fecha = obtener_fecha_exif(src)
        if fecha:
            imagenes_con_fecha.append((src, fecha))
    
    if not imagenes_con_fecha:
        return [], []

    imagenes_con_fecha.sort(key=lambda x: x[1])

    dia_x = imagenes_con_fecha[0][1].date()
    dia_z = imagenes_con_fecha[-1][1].date()

    inst = []
    desc = []

    for src, fecha in imagenes_con_fecha:
        if fecha.date() == dia_x:
            inst.append(src)
        else:
            desc.append(src)

    return inst, desc


def insert_image(ws, cellNumber, image_source):

    # === 1) Preparar la imagen PIL ===
    if isinstance(image_source, PImage.Image):
        pil_img = image_source

    elif isinstance(image_source, (str, Path)):
        pil_img = PImage.open(str(image_source))

    else:
        # asumimos BytesIO
        image_source.seek(0)
        pil_img = PImage.open(image_source)

    pil_img.load()  # asegurar que está cargada en memoria


    # Convertir PIL → openpyxl.Image 
    # openpyxl NO acepta PIL directamente, necesita un archivo o BytesIO
    temp_bytes = BytesIO()
    pil_img.save(temp_bytes, format="PNG")
    temp_bytes.seek(0)

    img = Image(temp_bytes)  # esta es la imagen de openpyxl


    #  Obtener info de la celda
    cell = ws[cellNumber]

    col_letter = cell.column_letter
    row_idx = cell.row

    col_dim = ws.column_dimensions.get(col_letter)
    row_dim = ws.row_dimensions.get(row_idx)

    col_width = col_dim.width if (col_dim and col_dim.width) else 8
    row_height = row_dim.height if (row_dim and row_dim.height) else 15

    max_width_px = col_width * 8
    max_height_px = row_height * 1.33


    #  Calcular escala
    if pil_img.width == 0 or pil_img.height == 0:
        ratio = 1.0
    else:
        ratio = min(max_width_px / pil_img.width,
                    max_height_px / pil_img.height,
                    1.0)

    img.width = int(pil_img.width * ratio)
    img.height = int(pil_img.height * ratio)


    # Insertar en Excel
    ws.add_image(img, cellNumber)
    
'''
def insert_image(ws, cellNumber, image_source):
    if isinstance(image_source, (str, Path)):
        img = Image(str(image_source))
    else:
        # asume BytesIO
        image_source.seek(0)
        img = Image(image_source)

    cell = ws[cellNumber]

    # Valores por defecto si no existen en la hoja
    # column_dimensions uses letters, row_dimensions uses numeric index
    col_letter = cell.column_letter
    row_idx = cell.row

    col_dim = ws.column_dimensions.get(col_letter)
    row_dim = ws.row_dimensions.get(row_idx)

    # ancho columna en "unidades Excel" -> aproximamos píxeles multiplicando por ~8
    col_width = col_dim.width if (col_dim and col_dim.width) else 8
    row_height = row_dim.height if (row_dim and row_dim.height) else 15  # puntos

    # Conversión aproximada (empírica): 1 unidad de columna ~ 8 píxeles, 1 punto ~ 1.33 px
    max_width_px = col_width * 8
    max_height_px = row_height * 1.33

    # Evitar división por 0
    if img.width == 0 or img.height == 0:
        ratio = 1.0
    else:
        ratio = min(max_width_px / img.width, max_height_px / img.height, 1.0)

    img.width = int(img.width * ratio)
    img.height = int(img.height * ratio)

    ws.add_image(img, cellNumber)
'''

def normalizar_booleans(data):
    if isinstance(data, dict):
        return {k: normalizar_booleans(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [normalizar_booleans(v) for v in data]
    elif isinstance(data, bool):
        return "Si" if data else "No"
    elif isinstance(data, str):
        if data.lower() == "true": return "Si"
        if data.lower() == "false": return "No"
    elif data is None:
        return ""
    return data

def convertir_valores_fecha(data):
    def convertir(k, v):
        if k == "CAMPO_EXTRA_8_TEXT" and isinstance(v, str):
            try:
                dt = datetime.fromisoformat(v)  # Parsear fecha ISO8601
                dt = dt.replace(tzinfo=None)   # Quitar timezone
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pass


        if "fecha" in k.lower():
            try:
                if isinstance(v, (int, float)) or (isinstance(v, str) and v.isdigit()):
                    ts = int(v) / 1000
                    dt = datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(timezone(timedelta(hours=-5)))
                    return dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
        return v
    if isinstance(data, dict):
        return {k: convertir(k, convertir_valores_fecha(v)) for k, v in data.items()}
    elif isinstance(data, list):
        return [convertir_valores_fecha(v) for v in data]
    return data

def obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, CIRCUITO_ACU):
    prefix = f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
    if "Contents" not in response:
        return {}
    json_files = [obj for obj in response["Contents"] if obj["Key"].lower().endswith(".json")]
    if not json_files:
        return {}

    latest_json = max(json_files, key=lambda x: x["LastModified"])["Key"]
    tmp_path = TMP_DIR / f"capa_principal_{GlobalID}.json"
    s3.download_file(bucket_name, latest_json, str(tmp_path))
    with open(tmp_path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except:
            return {}

def generar_hoja_desde_template(
    wb_final,
    payload_data,
    capa_principal_data,
    tipo_punto,
    imagen_keys,
    bucket_name,
    template_path_s3,
    template_name, code_key):
    """
    Genera una hoja nueva dentro del workbook consolidado usando la plantilla
    correspondiente al tipo de punto. Mantiene formato, merges, imágenes,
    bordes, fuentes y tamaños de celda, e inserta imágenes (tanto attachments como
    imágenes de Capa_principal).
    """

    #tipo_key = tipo_punto.lower()
    template_filename = template_name.get(code_key)
    if not template_filename:
        raise ValueError(f"No se encontró plantilla para tipo de punto: {tipo_punto}")

    template_key = template_path_s3 + template_filename

    # Descargar plantilla desde S3 a BytesIO
    template_stream = BytesIO()
    s3.download_fileobj(bucket_name, template_key, template_stream)
    template_stream.seek(0)

    # Cargar workbook de plantilla
    wb_template = load_workbook(template_stream)
    ws_template = wb_template.active

    # Obtener FID (convertir a string)
    fid = str(payload_data.get("FID_ELEM", "SIN_FID"))+'-RDP'
    sheet_name = fid[:31] if fid else "SIN_FID"
    ws_new = wb_final.create_sheet(title=sheet_name)

    # === Copiar celdas y estilos ===
    for row_idx, row in enumerate(ws_template.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = ws_new.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
                except Exception:
                    # en caso de que alguna copia falle, no rompemos la ejecución
                    pass

    # Copiar merges
    for merged_range in ws_template.merged_cells.ranges:
        ws_new.merge_cells(str(merged_range))

    # Copiar tamaños de columnas
    for col_letter, dim in ws_template.column_dimensions.items():
        if dim.width:
            ws_new.column_dimensions[col_letter].width = dim.width

    # Copiar alturas de filas
    for row_idx, dim in ws_template.row_dimensions.items():
        try:
            if dim.height:
                ws_new.row_dimensions[row_idx].height = dim.height
        except Exception:
            pass

    # === AJUSTE MANUAL DE CELDAS ===
    if tipo_punto in ["puntos_medicion", "vrp"]:
        for col in ["B", "C", "D", "E"]:
            ws_new.column_dimensions[col].width = 40  # ancho ideal para imágenes

    # === Copiar formato condicional ===
    cf_template = ws_template.conditional_formatting
    for cf_range, rules in getattr(cf_template, "_cf_rules", {}).items():
        for rule in rules:
            try:
                ws_new.conditional_formatting.add(cf_range, rule)
            except Exception as e:
                print(f"No se pudo copiar formato condicional para {cf_range}: {e}")       


    # === Insertar imágenes propias de la plantilla (logos, etc.) ===
    # openpyxl stores images as Image objects in _images; extraemos su data si es posible.
    for img in getattr(ws_template, "_images", []):
        try:
            # intentamos obtener bytes si el objeto lo contiene
            img_bytes = None
            # si es Image cargada de archivo local, openpyxl Image tiene .ref o .path en algunas versiones
            try:
                img_bytes = BytesIO(img._data())
            except Exception:
                # fallback: si es Image referenciada por path
                try:
                    img_bytes = BytesIO(open(img.ref, "rb").read())
                except Exception:
                    img_bytes = None

            if img_bytes:
                new_img = Image(img_bytes)
                # Mantener ancla aproximada
                try:
                    new_img.anchor = img.anchor
                except Exception:
                    pass
                ws_new.add_image(new_img)
        except Exception:
            # no rompemos por una imagen de plantilla
            pass

    payload_data = convertir_valores_fecha(payload_data)
    capa_principal_data = convertir_valores_fecha(capa_principal_data)

    # === Reemplazar placeholders con datos ===
    def reemplazar_valor(texto):
        if not isinstance(texto, str):
            return texto
        pattern = re.compile(r"\{\{(.*?)\}\}")
        def repl(match):
            key = match.group(1).strip()
            # Buscar en payload primero, luego en capa principal
            if key in payload_data:
                return str(payload_data[key] if payload_data[key] is not None else "")
            if key in capa_principal_data:
                return str(capa_principal_data[key] if capa_principal_data[key] is not None else "")
            return match.group(0)
        return pattern.sub(repl, texto)

    for row in ws_new.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = reemplazar_valor(cell.value)

    # === Preparamos lista de celdas disponibles para imágenes según plantilla ===
    celdas_imagenes = celdas_imagenes_plantilla.get(code_key, [])

    # === Descargar imágenes (attachments + capa_principal) localmente/stream y luego insertarlas ===
    # imagen_keys puede contener rutas locales (files/...) o llaves S3. Detectamos S3 por que no empiezan con '/' ni 'tmp' ni 'files/temp-image-folder'
    imagen_streams = []
    for key in imagen_keys:
        # evitar None o strings vacíos
        if not key or not isinstance(key, str):
            continue
        # En algunos payloads attachments vienen concatenados a un string; intentar separar por comas si detectamos eso
        if "," in key and (key.startswith("CW") or key.startswith("files/")):
            # si es un string con varias keys separadas por comas, separar
            parts = [p.strip() for p in key.split(",") if p.strip()]
        else:
            parts = [key.strip()]

        for k in parts:
            # Si es ya una ruta local (por ejemplo 'files/temp-image-folder/ejemplo1.jpg') o startswith('/') usamos directo
            if k.startswith("/") or k.startswith("files/") or k.startswith("tmp/"):
                local = Path(TMP_DIR) / Path(k).name
                # si no existe, intentar descargar si parece S3 key (empieza por files/ suele ser key en S3)
                if not local.exists():
                    # intentar descargar desde S3 si la key existe
                    try:
                        s3.download_file(bucket_name, k, str(local))
                    except Exception:
                        # no se pudo descargar, saltar
                        continue
                imagen_streams.append(local)
            else:
                # Probablemente es una key de S3 (p.e. "ArcGIS-Data/...")
                local = TMP_DIR / Path(k).name
                try:
                    # si la key incluye prefijo tipo "CW357221-ArcGIS-Data/..." y viene con coma, limpiamos prefijo accidental
                    # A veces attachments traen "CW357221-<key>", si detectamos '-' antes de 'ArcGIS-Data' intentamos extraer la parte desde 'ArcGIS-Data'
                    if "-" in k and "ArcGIS-Data" in k:
                        parts_dash = k.split("-", 1)
                        candidate = parts_dash[1]
                        # si candidate existe en s3, preferirlo
                        try:
                            s3.head_object(Bucket=bucket_name, Key=candidate)
                            k_to_use = candidate
                        except Exception:
                            k_to_use = k
                    else:
                        k_to_use = k

                    s3.download_file(bucket_name, k_to_use, str(local))
                    imagen_streams.append(local)
                except Exception:
                    # como fallback intentar descargar directamente con la key original
                    try:
                        s3.download_file(bucket_name, k, str(local))
                        imagen_streams.append(local)
                    except Exception:
                        # no pudimos descargar la imagen, la ignoramos
                        continue

    # === Insertar imágenes en las celdas asignadas ===
    '''
    for celda, img_src in zip(celdas_imagenes, imagen_streams):
        try:
            insert_image(ws_new, celda, img_src)
        except Exception:
            # si falla una imagen, seguimos con las siguientes
            continue
    '''
    # Clasificar imágenes según EXIF (inst / desc)
    inst_imgs, desc_imgs = clasificar_imagenes_por_fecha(imagen_streams)

    # Obtener celdas desde la plantilla
    celdas = celdas_imagenes_plantilla.get(code_key, {})
    celdas_inst = celdas.get("inst", [])
    celdas_desc = celdas.get("desc", [])

    # Insertar imágenes de instalación
    for celda, img_src in zip(celdas_inst, inst_imgs):
        try:
            insert_image(ws_new, celda, img_src)
        except Exception:
            continue

    # Insertar imágenes de desinstalación
    for celda, img_src in zip(celdas_desc, desc_imgs):
        try:
            insert_image(ws_new, celda, img_src)
        except Exception:
            continue



    return ws_new


def generar_resumen_desde_template(wb_final, lista_puntos, datos_globales, bucket_name, template_path_s3):
    """
    Genera la hoja 'Resumen' copiando desde la plantilla,
    incluyendo estilos, merges, imágenes, tamaños, formato condicional,
    y replicando una sola fila con placeholders en repeat.
    """

    template_filename = "listado-senales.xlsx"
    template_key = f"{template_path_s3}{template_filename}"

    # === Descargar plantilla ===
    template_stream = BytesIO()
    s3.download_fileobj(bucket_name, template_key, template_stream)
    template_stream.seek(0)

    wb_template = load_workbook(template_stream)
    ws_template = wb_template.active

    # === Crear hoja resultado ===
    ws_resumen = wb_final.create_sheet("LSE")

    #Para que quede al inicio
    wb_final._sheets.remove(ws_resumen)
    wb_final._sheets.insert(0, ws_resumen)

    # === Copiar celdas + estilos completos ===
    for row_idx, row in enumerate(ws_template.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):

            new_cell = ws_resumen.cell(row=row_idx, column=col_idx, value=cell.value)

            if cell.has_style:
                try:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
                except:
                    pass

    # === Copiar merges ===
    for merged_range in ws_template.merged_cells.ranges:
        ws_resumen.merge_cells(str(merged_range))

    # === Copiar tamaños de columnas ===
    for col_letter, dim in ws_template.column_dimensions.items():
        try:
            if dim.width:
                ws_resumen.column_dimensions[col_letter].width = dim.width
        except:
            pass

    # === Copiar alturas de filas ===
    for idx, dim in ws_template.row_dimensions.items():
        try:
            if dim.height:
                ws_resumen.row_dimensions[idx].height = dim.height
        except:
            pass

    # === AJUSTE MANUAL DE CELDAS ===
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws_resumen.column_dimensions[col].width = 22  # ancho ideal            

    # === Copiar formato condicional ===
    try:
        for cf_range, rules in getattr(ws_template.conditional_formatting, "_cf_rules", {}).items():
            for rule in rules:
                ws_resumen.conditional_formatting.add(cf_range, rule)
    except Exception as e:
        print(f"Error copiando formato condicional: {e}")

    # === Copiar imágenes (logos, etc.) ===
    for img in getattr(ws_template, "_images", []):
        try:
            img_bytes = None

            try:
                img_bytes = BytesIO(img._data())
            except:
                try:
                    img_bytes = BytesIO(open(img.ref, "rb").read())
                except:
                    img_bytes = None

            if img_bytes:
                new_img = Image(img_bytes)
                try:
                    new_img.anchor = img.anchor
                except:
                    pass
                ws_resumen.add_image(new_img)

        except:
            pass
  
    #placeholders globales
    for row in ws_resumen.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                val = cell.value
                for k, v in datos_globales.items():
                    val = val.replace(f"{{{{global.{k}}}}}", str(v if v is not None else ""))
                cell.value = val

    # === Buscar la fila plantilla con placeholders ===
    fila_template = None
    for r in range(1, ws_template.max_row + 1):
        for cell in ws_template[r]:
            if isinstance(cell.value, str) and cell.value.strip() == "{{repeat:punto}}":
                fila_template = r
                break
        if fila_template:
            break

    if not fila_template:
        print("⚠ No se encontró {{repeat:punto}} en la plantilla.")
        return ws_resumen

    # Guardar estructura de la fila plantilla
    base_row = []
    for cell in ws_template[fila_template]:
        base_row.append({
            "value": cell.value,
            "font": cell.font.copy() if cell.has_style else None,
            "border": cell.border.copy() if cell.has_style else None,
            "fill": cell.fill.copy() if cell.has_style else None,
            "number_format": cell.number_format,
            "alignment": cell.alignment.copy() if cell.has_style else None,
        })


    # Eliminar la fila plantilla original
    ws_resumen.delete_rows(fila_template)

    current_row = fila_template

    for punto in lista_puntos:
        ws_resumen.insert_rows(current_row)

        for c, cell_tpl in enumerate(base_row, start=1):
            val = cell_tpl["value"]

            # Reemplazar valores
            if isinstance(val, str):
                if val.strip() == "{{repeat:punto}}":
                    val = punto.get("FID_ELEM", "")
                else:
                    for k, v in punto.items():
                        val = val.replace(f"{{{{{k}}}}}", str(v if v is not None else ""))
                    # Reemplazar global.xxx
                    for k, v in datos_globales.items():
                        val = val.replace(f"{{{{global.{k}}}}}", str(v if v is not None else ""))
                    

            new = ws_resumen.cell(row=current_row, column=c, value=val)

            if cell_tpl["font"]: new.font = cell_tpl["font"]
            if cell_tpl["border"]: new.border = cell_tpl["border"]
            if cell_tpl["fill"]: new.fill = cell_tpl["fill"]
            new.number_format = cell_tpl["number_format"]
            if cell_tpl["alignment"]: new.alignment = cell_tpl["alignment"]

        current_row += 1

    return ws_resumen


def agregar_registro_lse(capa_principal_data, payload_data, listado_senales, REGLAS, RANGOS_OPERACIONALES):
    payload_data = convertir_valores_fecha(payload_data)
    capa_principal_data = convertir_valores_fecha(capa_principal_data)
    base = {
        "FID_ELEM": capa_principal_data.get("FID_ELEM"),
        "TIPO_PUNTO": capa_principal_data.get("TIPO_PUNTO"),
        "OBSERV_ACU": capa_principal_data.get("OBSERV_ACU"),
        "CRITERIO_ACU": capa_principal_data.get("CRITERIO_ACU"),
        "VARIABLE_A_MEDIR": capa_principal_data.get("VARIABLE_A_MEDIR"),
        "CIRCUITO_ACU" : payload_data.get("CIRCUITO_ACU", ""),
        "FECHA_FASE3" : payload_data.get("FECHA_FASE3", ""),
        "CAMPO_EXTRA_8_TEXT" : payload_data.get("CAMPO_EXTRA_8_TEXT", ""),
    }

    tipo_punto = base["TIPO_PUNTO"]
    equipo = payload_data.get("EQUIPO__DATALOGGER_INSTALADOS", "")
    variable_medir = base["VARIABLE_A_MEDIR"]

    # siempre agregamos la fila base
    #listado_senales.append(base)

    for regla in REGLAS:
        cond = regla["condicion"]

        if (
            tipo_punto == cond.get("TIPO_PUNTO") and
            equipo == cond.get("EQUIPO__DATALOGGER_INSTALADOS") and
            variable_medir == cond.get("VARIABLE_A_MEDIR")
        ):
            #aplicar reglas
            for salida in regla["salidas"]:
                nuevo = base.copy()

                # agregar campos estáticos y dinámicos
                for k, v in salida.items():
                    if k == "identificador_datalogger":
                        nuevo[k] = payload_data.get(v, "")
                    else:
                        nuevo[k] = v

                #aplicar nuevo rango operacional
                if equipo.lower() == "additel":
                    identificador = nuevo.get("identificador_datalogger", "")
                    rango_dinamico = RANGOS_OPERACIONALES.get(identificador)

                    if rango_dinamico:
                        nuevo["rango_operacional"] = rango_dinamico
                    else:
                        # fallback a lo definido en la regla, si existe
                        if "rango_operacional" in salida:
                            nuevo["rango_operacional"] = salida["rango_operacional"]
                else:
                    # si no es Additel, solo usa lo definido en salida
                    if "rango_operacional" in salida:
                        nuevo["rango_operacional"] = salida["rango_operacional"]

                # agregar el registro generado
                listado_senales.append(nuevo)

            break  

# === Lambda Handler principal que genera un workbook con una hoja por punto ===
def lambda_handler(event, context):
    print(event)
    cod = event["payload"].get("COD", "")
    numero_consolidado = int(event["payload"].get("numero_consolidado", 0))
    circuito_acu = event["payload"].get("CIRCUITO_ACU", "").replace(" ", "_")
    FECHA_INICIO_MEDICION = event["payload"].get("FECHA_INICIO_MEDICION", "")
    FECHA_FIN_MEDICION = event["payload"].get("FECHA_FIN_MEDICION", "")

    # --- SISTEMA DE LOCK POR CIRCUITO ---
    lock_key = f"locks_formato/fase3/{circuito_acu}.lock"

    # Verificar si ya existe un lock para este circuito
    try:
        s3.put_object(
            Bucket=bucket_name,
            Key=lock_key,
            Body=b"LOCK",
            ContentType='text/plain',
            Metadata={'created': datetime.utcnow().isoformat()},
            IfNoneMatch='*'
        )
        print(f"Lock creado para {circuito_acu}")
    except ClientError as e:
        if e.response['Error']['Code'] == 'PreconditionFailed':
            print(f"Otro proceso ya tiene lock para {circuito_acu}")
            return {
                "statusCode": 200,
                "body": f"Lambda saltada: lock activo para {circuito_acu}"
            }
        raise    

    try:
        # === Buscar puntos con carpeta Fase3 ===
        prefix = f"ArcGIS-Data/Puntos/{circuito_acu}/"
        paginator = s3.get_paginator("list_objects_v2")

        subcarpetas_puntos = set()
        for page in paginator.paginate(Bucket=bucket_name, Prefix=prefix, Delimiter="/"):
            for c in page.get("CommonPrefixes", []):
                subcarpetas_puntos.add(c["Prefix"])

        puntos_fase3 = []
        for subfolder in subcarpetas_puntos:
            fase3_prefix = f"{subfolder}Fase3/"
            response = s3.list_objects_v2(Bucket=bucket_name, Prefix=fase3_prefix)
            if "Contents" not in response:
                continue
            json_files = [obj for obj in response["Contents"] if obj["Key"].lower().endswith(".json")]
            if not json_files:
                continue
            # Tomar el JSON más reciente de Fase3
            latest_json = max(json_files, key=lambda x: x["LastModified"])["Key"]
            puntos_fase3.append(latest_json)

        if not puntos_fase3:
            return {"status": "no_points_found", "circuito": circuito_acu}

        print(f"Se encontraron {len(puntos_fase3)} puntos Fase3 para {circuito_acu}")

        wb_final = Workbook()

        # Quitar la hoja creada por defecto para evitar interferencias
        ws_default = wb_final.active
        wb_final.remove(ws_default)
        #ws_default = wb_final.active
        #ws_default.title = "Resumen"

        listado_senales = []

        #cargar reglas para pagina LSE
        def cargar_json(json_name):
            ruta = os.path.join(os.path.dirname(__file__), json_name)
            with open(ruta, "r", encoding="utf-8") as f:
                return json.load(f)

        REGLAS = cargar_json("reglas_senales.json")
        RANGOS_OPERACIONALES = cargar_json("rangos_operacionales_additel.json")

        for json_key in puntos_fase3:
            tmp_json = TMP_DIR / f"punto_{Path(json_key).name}"
            s3.download_file(bucket_name, json_key, str(tmp_json))
            with open(tmp_json, "r", encoding="utf-8") as f:
                try:
                    payload_data = json.load(f)
                except Exception:
                    payload_data = {}

            tipo_punto = payload_data.get("TIPO_PUNTO", "").lower()
            GlobalID = payload_data.get("PARENT_ID")
            fid = payload_data.get("FID_ELEM", "SIN_FID")
            VARIABLES_MEDICION = payload_data.get("VARIABLES_MEDICION", "")
            EQUIPO__DATALOGGER_INSTALADOS = payload_data.get("EQUIPO__DATALOGGER_INSTALADOS", "")

            if tipo_punto == "vrp":
                code_key = tipo_punto + "-" + VARIABLES_MEDICION + "-" + EQUIPO__DATALOGGER_INSTALADOS
            else:
                code_key = tipo_punto

            print(code_key)

            # Obtener capa principal (atributos + posiblemente imágenes en esa carpeta)
            capa_principal_data = obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, circuito_acu)

            # Para pagina inicial "LSE"
            '''
            listado_senales.append({
                "FID_ELEM": capa_principal_data.get("FID_ELEM"),
                "VARIABLE_A_MEDIR": VARIABLE_A_MEDIR,
                "OBSERV_ACU": capa_principal_data.get("OBSERV_ACU"),
                "CRITERIO_ACU": capa_principal_data.get("CRITERIO_ACU")
            })'''
            agregar_registro_lse(capa_principal_data, payload_data, listado_senales, REGLAS, RANGOS_OPERACIONALES)

            # "traducir" valor de variable a medir en capa principal
            raw_var = capa_principal_data.get("VARIABLE_A_MEDIR", "")
            VARIABLE_A_MEDIR = variables_a_medir_traduccion.get(raw_var, raw_var)

            # reemplazar tambien en la key de payload data
            payload_data["VARIABLES_MEDICION"] = VARIABLE_A_MEDIR

            # reemplazar en la key equipo datalogger instalado
            payload_data["EQUIPO__DATALOGGER_INSTALADOS"] = datalogger_traduccion.get(payload_data.get("EQUIPO__DATALOGGER_INSTALADOS", ""))

            # Buscar imágenes dentro de Fase3
            folder = f"ArcGIS-Data/Puntos/{circuito_acu}/{GlobalID}_{tipo_punto}/Fase3/"
            resp = s3.list_objects_v2(Bucket=bucket_name, Prefix=folder)
            imagen_keys = []
            if "Contents" in resp:
                imagen_keys = [x["Key"] for x in resp["Contents"] if x["Key"].lower().endswith((".jpg", ".jpeg", ".png"))]

            
            #Ya no se necesitan las imagenes de capa principal
            '''
            folder_cp = f"ArcGIS-Data/Puntos/{circuito_acu}/{GlobalID}_{tipo_punto}/Capa_principal/"
            resp_cp = s3.list_objects_v2(Bucket=bucket_name, Prefix=folder_cp)
            if "Contents" in resp_cp:
                imagen_keys_cp = [x["Key"] for x in resp_cp["Contents"] if x["Key"].lower().endswith((".jpg", ".jpeg", ".png"))]
                # unir evitando duplicados
                for k in imagen_keys_cp:
                    if k not in imagen_keys:
                        imagen_keys.append(k)
            '''

            # También soportar attachments que vengan dentro del JSON payload (campo attachments o attachments list)
            attachments = []
            if isinstance(payload_data.get("attachments"), list):
                attachments = payload_data.get("attachments")
            elif isinstance(payload_data.get("attachments"), str):
                # si vienen como "a,b,c" intentar separar
                attachments = [p.strip() for p in payload_data.get("attachments").split(",") if p.strip()]

            # unir attachments (siempre como S3 keys o rutas)
            for a in attachments:
                if a not in imagen_keys:
                    imagen_keys.append(a)

            # Normalmente generar_hoja_desde_template descargará las imágenes desde S3 según las keys
            generar_hoja_desde_template(
                wb_final,
                payload_data,
                capa_principal_data,
                tipo_punto,
                imagen_keys,
                bucket_name,
                template_path_s3,
                template_name, code_key
            )
        print("listado señales", listado_senales)
       
        
        utc_minus_5 = timezone(timedelta(hours=-5))
        fecha_actual = datetime.now(utc_minus_5)
        fecha_reporte = fecha_actual.strftime("%Y-%m-%d")

        datos_globales = {
            "CIRCUITO_ACU": circuito_acu.replace("_", " "),
            "FECHA_CONSOLIDADO": fecha_reporte,
            "FECHA_INICIO_MEDICION" : FECHA_INICIO_MEDICION,
            "FECHA_FIN_MEDICION" : FECHA_FIN_MEDICION
        }
        print(datos_globales)
        generar_resumen_desde_template(wb_final,listado_senales, datos_globales, bucket_name,template_path_s3)

        #if ws_resumen.max_row == 1 and ws_resumen.max_column == 1:
        #    # si no quedó contenido en hoja resumen, borrarla
        #    try:
        #        wb_final.remove(ws_default)
        #    except Exception:
        #        pass

        output_filename = f"MPH-EJ-0601-{cod}-F03-ACU-LSE-001.xlsx"
        output_key = f"{output_path_s3}ACU/CIR/{output_filename}"
        output_path = TMP_DIR / output_filename
        wb_final.save(output_path)
        s3.upload_file(str(output_path), bucket_name, output_key)

        return {
            "status": "ok",
            "circuito": circuito_acu,
            "puntos_total": len(puntos_fase3),
            "output_file": f"s3://{bucket_name}/{output_key}"
        }

    finally:
        # --- Liberar el lock ---
        try:
            s3.delete_object(Bucket=bucket_name, Key=lock_key)
            print(f"Lock liberado para {circuito_acu}")
        except Exception as e:
            print(f"Error al eliminar lock para {circuito_acu}: {e}")