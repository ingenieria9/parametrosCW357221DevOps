import boto3
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image  # para insertar imágenes
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.utils import column_index_from_string
import os
from datetime import datetime, timezone, timedelta
import re

from PIL import Image as PImage, ExifTags


s3 = boto3.client("s3")
client_lambda_db = boto3.client("lambda", region_name="us-east-1") 
client_lambda = boto3.client("lambda") 
TMP_DIR = Path("/tmp")

# Parámetros de entrada (variables)

bucket_name = os.environ['BUCKET_NAME']
db_access_arn = os.environ['DB_ACCESS_LAMBDA_ARN']
FORMATO_CONSOLIDADO_LAMBDA_ARN = os.environ['FORMATO_CONSOLIDADO_LAMBDA_ARN']
template_path_s3 = "files/plantillas/Fase3/"
output_path_s3 = "files/entregables/Fase3/"
output_path_s3_for_convert = "files/files-to-convert/Fase3/"


template_name = {"puntos_medicion": "formato-acueducto-pm.xlsx",
                 "vrp-caudal-PLUM": "formato-acueducto-vrp-caudal-PLUM.xlsx",
                 "vrp-presion_caudal-PLUM": "formato-acueducto-vrp-presion_caudal-PLUM.xlsx",
                 "vrp-presion-Additel": "formato-acueducto-vrp-presion-Additel.xlsx",
                  "vrp-presion-PLUM":  "formato-acueducto-vrp-presion-PLUM.xlsx",
                   "camara": "formato-alcantarillado.xlsx"}

#MPH-EJ-0601-{CIR_COD}-F01-{ACU/ALC}-EIN-{FID}
COD_name = {"puntos_medicion": "ACU/PM/MPH-EJ-0601-{COD}-F03-ACU-RDP-",
            "vrp": "ACU/VRP/MPH-EJ-0601-{COD}-F03-ACU-RDP-", "camara": "ALC/MPH-EJ-0601-{COD}-F03-ALC-RDP-"}

'''celdas_imagenes_plantilla = {"puntos_medicion": ["B22", "C22", "D22", "E22","B23", "C23", "D23", "E23", "B24", "C24", "D24", "E24"],
                            "vrp-caudal-PLUM": ["B22", "C22", "D22", "E22","B23", "C23", "D23", "E23", "B24", "C24", "D24", "E24"],
                            "vrp-presion_caudal-PLUM": ["B24", "C24", "D24", "E24","B25", "C25", "D25", "E25", "B26", "C26", "D26", "E26"],
                            "vrp-presion-Additel": ["B27", "C27", "D27", "E27","B28", "C28", "D28", "E28", "B29", "C29", "D29", "E29"],
                            "vrp-presion-PLUM": ["B24", "C24", "D24", "E24","B25", "C25", "D25", "E25", "B26", "C26", "D26", "E26"], "camara": []}
'''
celdas_imagenes_plantilla = {
    "vrp-presion-Additel": {
        "inst": ["B33", "C33", "D33", "B34", "C34", "D34", "B35", "C35", "D35"],
        "desc": ["B37", "C37", "D37", "B38", "C38", "D38", "B39", "C39", "D39"]
    },
    "vrp-presion-PLUM": {
        "inst": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"],
        "desc": ["B33", "C33", "D33",  "B34", "C34", "D34", "B35", "C35", "D35"]
    },
    "vrp-presion_caudal-PLUM": {
        "inst": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"],
        "desc": ["B33", "C33", "D33"  "B34", "C34", "D34", "B35", "C35", "D35"]
    },
    "vrp-caudal-PLUM": {
        "inst": ["B25", "C25", "D25", "B26", "C26", "D26", "B27", "C27", "D27"],
        "desc": ["B29", "C29", "D29", "B30", "C30", "D30", "B31", "C31", "D31"]
    },
    "puntos_medicion": {
        "inst": ["B26", "C26", "D26", "B27", "C27", "D27", "B28", "C28", "D28"],
        "desc": ["B30", "C30", "D30", "B31", "C31", "D31", "B32", "C32", "D32"]
    },
    "camara": { #to be defined
        "inst": ["B24", "C24", "D24"],
        "desc": ["B28", "C28", "D28"]
    }
}


variables_a_medir_traduccion = {
    "presion_caudal": "Presión, Caudal","presion": "Presión",
    "caudal": "Caudal","area_velocidad": "Área, velocidad"}                        


def obtener_fecha_exif(imagen_path):
    try:
        img = PImage.open(imagen_path)
        exif_data = img._getexif()
        if not exif_data:
            return None

        exif = {
            ExifTags.TAGS.get(tag, tag): value
            for tag, value in exif_data.items()
        }

        fecha_raw = exif.get("DateTimeOriginal") or exif.get("DateTime")
        if not fecha_raw:
            return None
        print(fecha_raw)
        return datetime.strptime(fecha_raw, "%Y:%m:%d %H:%M:%S")
    
    except Exception:
        return None

def clasificar_imagenes_por_fecha(imagen_paths):
    imagenes_con_fecha = []

    for path in imagen_paths:
        fecha = obtener_fecha_exif(path)
        if fecha is None:
            continue
        imagenes_con_fecha.append((path, fecha))
    
    if not imagenes_con_fecha:
        return [], []

    # Ordenar por fecha ascendente
    imagenes_con_fecha.sort(key=lambda x: x[1])

    # Día X = instalación (más antiguo)
    dia_x = imagenes_con_fecha[0][1].date()
    # Día Z = desinstalación (más reciente)
    dia_z = imagenes_con_fecha[-1][1].date()

    inst = []
    desc = []

    for path, fecha in imagenes_con_fecha:
        if fecha.date() == dia_x:
            inst.append(path)
        else:
            desc.append(path)

    return inst, desc        



def insert_image(ws, cellNumber, imagen_path):

    # Abrir imagen con PIL para poder leer dimensiones reales
    pil_img = PImage.open(str(imagen_path))

    # Crear imagen para openpyxl
    img = Image(str(imagen_path))

    cell = ws[cellNumber]

    # Medidas de la celda
    col_width = ws.column_dimensions[cell.column_letter].width or 8
    row_height = ws.row_dimensions[cell.row].height or 15

    # Conversión aproximada a píxeles
    max_width = col_width * 8
    max_height = row_height * 1.0

    # Escala manteniendo proporciones (usamos dimensiones reales de PIL)
    ratio = min(max_width / pil_img.width, max_height / pil_img.height)

    img.width = pil_img.width * ratio
    img.height = pil_img.height * ratio

    ws.add_image(img, cellNumber)

def normalizar_booleans(data_get):
    def convertir_valor(valor):
        if valor is None:  # convertir None a cadena vacía
            return ""
        if isinstance(valor, bool):  # True/False nativos de Python
            return "Si" if valor else "No"
        if isinstance(valor, str):  # "true"/"false" como string
            if valor.lower() == "true":
                return "Si"
            if valor.lower() == "false":
                return "No"
        return valor  # cualquier otro valor queda igual

    # Si es un dict, lo recorremos recursivamente
    if isinstance(data_get, dict):
        return {k: normalizar_booleans(v) for k, v in data_get.items()}
    # Si es una lista, también recursivamente
    if isinstance(data_get, list):
        return [normalizar_booleans(v) for v in data_get]
    
    return convertir_valor(data_get)

def convertir_valores_fecha(data):
    """
    Convierte valores tipo timestamp (en milisegundos) a formato legible
    solo si la clave contiene la palabra 'fecha' (insensible a mayúsculas).
    Interpreta el timestamp como UTC y lo convierte a UTC-5.
    Funciona de forma recursiva para dicts y listas.
    """

    def convertir_fecha(key, valor):
        try:
            if "fecha" in key.lower() or "CAMPO_EXTRA_8_TEXT" in key:
                if isinstance(valor, (int, float)) or (isinstance(valor, str) and valor.isdigit()):
                    timestamp = int(valor) / 1000  # convertir a segundos
                    # Interpretar en UTC y convertir a UTC-5
                    dt_utc = datetime.fromtimestamp(timestamp, tz=timezone.utc)
                    dt_utc_minus_5 = dt_utc.astimezone(timezone(timedelta(hours=-5)))
                    return dt_utc_minus_5.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(f"Error al convertir fecha ({key}): {e}")
        return valor

    if isinstance(data, dict):
        nuevo_dict = {}
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                nuevo_dict[k] = convertir_valores_fecha(v)
            else:
                nuevo_dict[k] = convertir_fecha(k, v)
        return nuevo_dict

    elif isinstance(data, list):
        return [convertir_valores_fecha(v) for v in data]

    else:
        return data

def obtener_consecutivo_s3(bucket, prefix, cod_name):
    """
    Busca el mayor número consecutivo en los nombres de archivos dentro del prefijo dado
    y devuelve el siguiente número disponible (formato 001, 002, etc.).
    """
    max_consec = 0
    pattern = re.escape(cod_name) + r"(\d{3})\.xlsx$"

    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            match = re.search(pattern, key)
            if match:
                num = int(match.group(1))
                if num > max_consec:
                    max_consec = num

    return f"{max_consec + 1:03}"


    # Tomar el más reciente por fecha
    latest_json = max(json_files, key=lambda x: x["LastModified"])["Key"]

    print(f"Usando archivo principal: {latest_json}")

    # Descargar temporalmente en /tmp
    tmp_path = TMP_DIR / "capa_principal.json"
    s3.download_file(bucket_name, latest_json, str(tmp_path))

    # Leer el contenido con validación
    with open(tmp_path, "r", encoding="utf-8") as f:
        contenido = f.read().strip()
        if not contenido:
            print(" El archivo JSON está vacío.")
            return {}
        try:
            return json.loads(contenido)
        except json.JSONDecodeError as e:
            print(f" Error al parsear JSON {latest_json}: {e}")
            return {}
        
def invoke_lambda_db(payload, FunctionName):
    response = client_lambda_db.invoke(
        FunctionName=FunctionName,
        InvocationType='RequestResponse',
        Payload=json.dumps(payload).encode('utf-8')
    )
    
    # Lee el cuerpo de la respuesta
    result = response["Payload"].read().decode("utf-8")
    
    # Intenta parsear a JSON si es posible
    try:
        return json.loads(result)
    except json.JSONDecodeError:
        return {"raw_response": result}      

def invoke_lambda(payload, FunctionName):
    response = client_lambda.invoke(
        FunctionName=FunctionName,
        InvocationType='Event',  # async
        Payload=json.dumps(payload).encode('utf-8')  #  convierte a JSON y luego a bytes
    )
    return response  


def obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, CIRCUITO_ACU):
    # Construir el prefijo correcto
    s3_key_capa_principal = (
        f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    )

    # Listar objetos en esa carpeta
    s3_objects = s3.list_objects_v2(Bucket=bucket_name, Prefix=s3_key_capa_principal)

    if "Contents" not in s3_objects:
        print("No hay archivos en la carpeta Capa_principal.")
        return {}

    # Filtrar SOLO archivos que terminen en .json
    json_files = [
        obj for obj in s3_objects["Contents"]
        if obj["Key"].lower().endswith(".json")
    ]

    if not json_files:
        print("No se encontraron archivos .json en Capa_principal.")
        print(f"Se buscó usando Prefix: {s3_key_capa_principal}")
        return {}

    # Tomar el más reciente por fecha
    latest_json = max(json_files, key=lambda x: x["LastModified"])["Key"]

    print(f"Usando archivo principal: {latest_json}")

    # Descargar temporalmente en /tmp
    tmp_path = TMP_DIR / "capa_principal.json"
    s3.download_file(bucket_name, latest_json, str(tmp_path))

    # Leer el contenido con validación
    with open(tmp_path, "r", encoding="utf-8") as f:
        contenido = f.read().strip()
        if not contenido:
            print(" El archivo JSON está vacío.")
            return {}
        try:
            return json.loads(contenido)
        except json.JSONDecodeError as e:
            print(f" Error al parsear JSON {latest_json}: {e}")
            return {}
        

def lambda_handler(event, context):

    payload_data = event["payload"]
    tipo_punto = event["payload"]["TIPO_PUNTO"]
    FID_ELEM = event["payload"]["FID_ELEM"]
    GlobalID = event["payload"]["PARENT_ID"] #id uuid global 
    CIRCUITO_ACU = event["payload"]["CIRCUITO_ACU"].replace(" ", "_")
    forzarInforme = event.get("forzarInforme", "false")

    #valores para armar key de la plantilla
    EQUIPO__DATALOGGER_INSTALADOS = event["payload"].get("EQUIPO__DATALOGGER_INSTALADOS", "")
    VARIABLES_MEDICION = event["payload"].get("VARIABLES_MEDICION", "")

    if tipo_punto == "vrp":
        value_code = tipo_punto + "-" + VARIABLES_MEDICION + "-" + EQUIPO__DATALOGGER_INSTALADOS
    else:
        value_code = tipo_punto

    # "traducir" valor de variable a medir 
    raw_var = payload_data.get("VARIABLES_MEDICION", "")
    VARIABLES_MEDICION = variables_a_medir_traduccion.get(raw_var, raw_var)

    # reemplazar tambien en la key de payload data
    payload_data["VARIABLES_MEDICION"] = VARIABLES_MEDICION        

    #template_path_s3 + devolver de template key el value segun tipo de punto y variables a medir
    template_key = template_path_s3 + template_name.get(value_code)

    # Paths locales en Lambda (/tmp)
    template_path = TMP_DIR / "plantilla.xlsx"
    imagen_keys = event["attachments"]

    # IMAGENES EXTRA DE CAPA PRINCIPAL (NO APLICA EN FASE 3)
    '''
    # otras imagenes extra de capa principal
    folder_imagen_extra = f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    # Listar objetos PNG en el folder del bucket
    response = s3.list_objects_v2(
        Bucket=bucket_name,
        Prefix=folder_imagen_extra
    )
    # Agregar todos los archivos .png encontrados
    if "Contents" in response:
        for obj in response["Contents"]:
            key = obj["Key"]
            if key.lower().endswith((".png", ".jpeg", ".jpg")):
                imagen_keys.append(key)
    # Eliminar duplicados si existen
    imagen_keys = list(set(imagen_keys))
    '''

    imagen_paths = [TMP_DIR / Path(k).name for k in imagen_keys]
    output_path = TMP_DIR / "output.xlsx"

    # Descargar archivos desde S3
    s3.download_file(bucket_name, template_key, str(template_path))
    #s3.download_file(bucket_name, json_key, str(json_path))
    
    #descargar imagenes
    for key, path in zip(imagen_keys, imagen_paths):
        print("descargando", key)
        s3.download_file(bucket_name, key, str(path))

    # Cargar plantilla Excel
    # Construir el prefijo correcto
    s3_key_capa_principal = (
        f"ArcGIS-Data/Puntos/{CIRCUITO_ACU}/{GlobalID}_{tipo_punto}/Capa_principal/"
    )

    # Listar objetos en esa carpeta
    s3_objects = s3.list_objects_v2(Bucket=bucket_name, Prefix=s3_key_capa_principal)

    if "Contents" not in s3_objects:
        print("No hay archivos en la carpeta Capa_principal.")
        return {}

    # Filtrar SOLO archivos que terminen en .json
    json_files = [
        obj for obj in s3_objects["Contents"]
        if obj["Key"].lower().endswith(".json")
    ]

    if not json_files:
        print("No se encontraron archivos .json en Capa_principal.")
        print(f"Se buscó usando Prefix: {s3_key_capa_principal}")
        return {}
    wb = load_workbook(template_path)
    ws = wb.active  # hoja específica con wb["NombreHoja"] o activa con wb.active

    # ajustar ancho columnas
    for col in ['B', 'C', 'D', 'E']:
        #print(f"Columna {col}: {ws.column_dimensions[col].width}")
        ws.column_dimensions[col].width = 40
        #print(f"Columna {col}: {ws.column_dimensions[col].width}")
    
    # Leer datos adicionales desde S3
    capa_principal_data = obtener_info_de_capa_principal(bucket_name, tipo_punto, GlobalID, CIRCUITO_ACU)
    #capa_principal_atributos = capa_principal_data["attributes"]
    # Unir ambos diccionarios (payload tiene prioridad si hay claves iguales)
    combined_data = {**capa_principal_data, **payload_data}

    #  Normalizar
    json_data = normalizar_booleans(combined_data)
    json_data = convertir_valores_fecha(json_data)

    # Campos para placeholders
    campos_contexto = list(combined_data.keys())

    # Construir contexto final
    context = {
        f"{{{{{campo}}}}}": json_data.get(campo, "")
        for campo in campos_contexto
    }

    print(context)


    # Reemplazar variables en todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for placeholder, value in context.items():
                    if placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, str(value))

    # INSERTAR IMAGENES
    celdas_tipo = celdas_imagenes_plantilla[value_code]
    print(celdas_tipo)

    # Clasificar imágenes por EXIF
    inst_imgs, desc_imgs = clasificar_imagenes_por_fecha(imagen_paths)

    print("inst_imgs", inst_imgs)
    print("desc_imgs",desc_imgs )

    # Inserción instalación
    for celda, img in zip(celdas_tipo["inst"], inst_imgs):
        insert_image(ws, celda, img)

    # Inserción desinstalación
    for celda, img in zip(celdas_tipo["desc"], desc_imgs):
        insert_image(ws, celda, img)

    #celdas_imagenes = celdas_imagenes_plantilla.get(value_code)
    #Ciclo para insertar las imagenes en las celdas disponibles
    #for celda, imagen_path in zip(celdas_imagenes, imagen_paths):
    #    insert_image(ws, celda, imagen_path)

    # Guardar archivo final
    wb.save(output_path)


    # Subir resultado a S3
    #listar los archivos de s3 en
    #Obtener el consecutivo siguiente en esa carpeta
    #consecutivo = obtener_consecutivo_s3(bucket_name, output_path_s3, COD_name[tipo_punto])

    #obtener de S3 el archivo json que contiene el codigo del circuito para construir el archivo
    # Descargar temporalmente en /tmp
    tmp_path_code = TMP_DIR / "code.json"
    if tipo_punto == "camara":
        code_file = "files/epm_codes/CODE_ALC_CUE.json"
    else:
        code_file = "files/epm_codes/CODE_ACU_CIR.json"
    s3.download_file(bucket_name, code_file, str(tmp_path_code))

    # Leer el contenido con validación
    with open(tmp_path_code, "r", encoding="utf-8") as f:
        contenido = f.read().strip()
        if not contenido:
            print(" El archivo JSON está vacío.")
        try:
            code_json =  json.loads(contenido)
        except json.JSONDecodeError as e:
            print(f" Error al parsear JSON {code_file}: {e}")
    
    code_data = code_json[event["payload"]["CIRCUITO_ACU"]]

    if not code_data:
        code_data = CIRCUITO_ACU

    file_name = COD_name[tipo_punto].format(COD=code_data)

    #Construir el nombre completo del archivo
    output_key = f"{output_path_s3}{file_name}{FID_ELEM}.xlsx"
    #convert_output_key = f"{output_path_s3_for_convert}{file_name}{FID_ELEM}.xlsx"

    #Subir a carpeta entregables y a files_to_convert
    #s3.upload_file(str(output_path), bucket_name, convert_output_key)
    s3.upload_file(str(output_path), bucket_name, output_key)


    # BUSCAR EN LA DB FECHA_INICIO_MEDICIONES Y FECHA_FIN_MEDICIONES
    payload_db = {
        "queryStringParameters": {
            "query": f"""SELECT "FECHA_INICIO_MEDICION" from fase_3_a_b_trazabilidad_mediciones fabtm where "CIRCUITO_ACU" = '{event["payload"]["CIRCUITO_ACU"]}'""",
            "time_column": "FECHA_INICIO_MEDICION",
            "db_name": "parametros"
        }
    }
    response_db =invoke_lambda_db(payload_db, db_access_arn)
    body = json.loads(response_db["body"])
    FECHA_INICIO_MEDICION = body[0]["FECHA_INICIO_MEDICION"]

    # ajustar utc
    dt = datetime.strptime(FECHA_INICIO_MEDICION, "%Y-%m-%d %H:%M:%S")
    # Ajustar a UTC-5 (restar 5 horas)
    dt = dt - timedelta(hours=5)
    # Convertir de nuevo a string si lo necesitas igual que antes
    FECHA_INICIO_MEDICION = dt.strftime("%Y-%m-%d %H:%M:%S")

    payload_db = {
        "queryStringParameters": {
            "query": f"""SELECT "FECHA_FIN_MEDICION" from fase_3_a_b_trazabilidad_mediciones fabtm where "CIRCUITO_ACU" = '{event["payload"]["CIRCUITO_ACU"]}'""",
            "time_column": "FECHA_FIN_MEDICION",
            "db_name": "parametros"
        }
    }
    response_db =invoke_lambda_db(payload_db, db_access_arn)
    body = json.loads(response_db["body"])
    FECHA_FIN_MEDICION = body[0]["FECHA_FIN_MEDICION"]

    # ajustar utc
    dt = datetime.strptime(FECHA_FIN_MEDICION, "%Y-%m-%d %H:%M:%S")
    # Ajustar a UTC-5 (restar 5 horas)
    dt = dt - timedelta(hours=5)
    # Convertir de nuevo a string si lo necesitas igual que antes
    FECHA_FIN_MEDICION = dt.strftime("%Y-%m-%d %H:%M:%S")

    print(FECHA_FIN_MEDICION)
    print(FECHA_INICIO_MEDICION)

    # TO-DO: Revisar si si es pertinente
    '''
    payload_db = {
        "queryStringParameters": {
            "query": f""" WITH circuito AS (SELECT "CIRCUITO_ACU" FROM puntos_capa_principal WHERE "GlobalID" = '{GlobalID}'),
            puntos_realizados AS (SELECT COUNT(*) AS count_realizados, p."CIRCUITO_ACU" FROM puntos_capa_principal p WHERE p."CIRCUITO_ACU" = (SELECT "CIRCUITO_ACU" FROM circuito) AND p."PUNTO_EXISTENTE" = 'Si' AND p."HABILITADO_FASE3" = 1 AND p."FID_ELEM" IN ( SELECT "FID_ELEM" FROM  fase_3_a_data where "El_punto_requiere_fase_3" = 'Si') GROUP BY p."CIRCUITO_ACU"),
            puntos_totales AS (SELECT  COUNT(*) AS count_totales,  p."CIRCUITO_ACU"  FROM puntos_capa_principal p WHERE  p."CIRCUITO_ACU" = ( SELECT    "CIRCUITO_ACU"  FROM circuito)  AND p."PUNTO_EXISTENTE" = 'Si' AND p."HABILITADO_FASE3" = 1 GROUP BY  p."CIRCUITO_ACU")
            SELECT CASE WHEN t.count_totales = r.count_realizados THEN 'Finalizado' ELSE 'Incompleto'  END AS estado,t."CIRCUITO_ACU", t.count_totales AS numero_puntos,  r.count_realizados AS puntos_realizados FROM puntos_totales t LEFT JOIN puntos_realizados r ON t."CIRCUITO_ACU" = r."CIRCUITO_ACU";""",
            "time_column": "fecha_creacion",
            "db_name": "parametros"
        }
    }
    print(payload_db)
    response_db =invoke_lambda_db(payload_db, db_access_arn)
    print(response_db)
    #Parsear el body 
    body = json.loads(response_db["body"])
    # Extraer el valor del campo "estado"
    estado = body[0]["estado"]
    circuito = body[0]["CIRCUITO_ACU"]
    numero_puntos = body[0]["numero_puntos"]
    #puntos_realizados = body[0]["puntos_realizados"]
    #puntos_realizados = body[0].get("puntos_realizados", 0)
    '''
    puntos_realizados = 0 #Temp value

    '''
    if estado == "Finalizado":   # Si es ultimo punto, invocar a lambda de generación de formato consolidado (async)
        incoming_payload = { "payload": { "COD": code_data, "numero_consolidado" : numero_puntos, "CIRCUITO_ACU" : CIRCUITO_ACU } }
        invoke_lambda(incoming_payload, FORMATO_CONSOLIDADO_LAMBDA_ARN)
        print("Invocada lambda formato consolidado")    
    '''
    
    if str(forzarInforme).lower() == "true": #si viene de la API Forzado a ejecutarse
        incoming_payload = { "payload": { "COD": code_data, "numero_consolidado" : puntos_realizados, "CIRCUITO_ACU" : CIRCUITO_ACU, "FECHA_FIN_MEDICION" : FECHA_FIN_MEDICION, "FECHA_INICIO_MEDICION" : FECHA_INICIO_MEDICION } } 
        print(incoming_payload)
        invoke_lambda(incoming_payload, FORMATO_CONSOLIDADO_LAMBDA_ARN)
        print("Invocada lambda formato consolidado")    
    return {
        "status": "ok",
        "output_file": f"s3://{bucket_name}/{output_key}"
    }